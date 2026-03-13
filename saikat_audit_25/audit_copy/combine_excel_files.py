#!/usr/bin/env python3
"""
Combine all Excel files from Expense details into a single CSV with receipt mapping
"""

import os
import csv
from pathlib import Path
from openpyxl import load_workbook
import sys
from difflib import SequenceMatcher
import pandas as pd

def load_receipt_mapping(mapping_csv):
    """Load the expense receipt mapping CSV into a dictionary for quick lookup"""
    mapping = {}
    try:
        with open(mapping_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                receipt_filename = row['Receipt_Filename'].strip()
                mapping[receipt_filename] = {
                    'event': row['Event'],
                    'excel_file': row['Excel_Filename'],
                    'full_path': row['Receipt_Full_Path'],
                    'row_number': reader.line_num - 1  # Row number in mapping CSV
                }
    except Exception as e:
        print(f"Error loading receipt mapping: {e}", file=sys.stderr)
        return {}
    return mapping

def fuzzy_match_filename(receipt_text, available_filenames, threshold=0.6):
    """
    Fuzzy match receipt text against available filenames
    Returns the best match and similarity score
    """
    if not receipt_text:
        return None, 0.0

    receipt_text = receipt_text.strip()
    best_match = None
    best_score = 0.0

    for filename in available_filenames:
        # Try exact match first
        if receipt_text == filename:
            return filename, 1.0

        # Try matching the end of the receipt text (as user mentioned)
        if receipt_text.endswith(filename):
            return filename, 1.0

        # Fuzzy match using sequence matcher
        score = SequenceMatcher(None, receipt_text.lower(), filename.lower()).ratio()
        if score > best_score and score >= threshold:
            best_match = filename
            best_score = score

        # Also try matching just the filename part if receipt_text has path separators
        if '/' in receipt_text or '\\' in receipt_text:
            basename = Path(receipt_text).name
            score = SequenceMatcher(None, basename.lower(), filename.lower()).ratio()
            if score > best_score and score >= threshold:
                best_match = filename
                best_score = score

    return best_match, best_score

def find_nested_folder(parent_path):
    """Find the nested folder with the same name pattern"""
    for item in parent_path.iterdir():
        if item.is_dir():
            return item
    return None

def find_all_xlsx_files(folder_path):
    """Find all Excel files in a folder"""
    xlsx_files = []
    for item in folder_path.iterdir():
        if item.is_file() and item.suffix.lower() in ['.xlsx', '.xls']:
            xlsx_files.append(item)
    return sorted(xlsx_files)

def process_excel_file(xlsx_path, subdir_name, receipt_mapping):
    """Process a single Excel file and return rows with mapping info"""
    rows = []
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Read header row
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if row and any(cell is not None for cell in row):
                header_row = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                break

        if not header_row:
            print(f"   ⚠️  No header row found in {xlsx_path.name}")
            return rows

        # Required columns (case-insensitive)
        required_cols = ['date', 'detail', 'type', 'amount', 'owner', 'receipt']
        col_indices = {}

        for req_col in required_cols:
            if req_col in header_row:
                col_indices[req_col] = header_row.index(req_col)
            else:
                print(f"   ⚠️  Required column '{req_col}' not found in {xlsx_path.name}")
                return rows  # Skip this file

        # Now process data rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue

            if not row or all(cell is None for cell in row):
                continue

            # Extract required columns
            date_val = row[col_indices['date']] if col_indices['date'] < len(row) else ''
            detail_val = row[col_indices['detail']] if col_indices['detail'] < len(row) else ''
            type_val = row[col_indices['type']] if col_indices['type'] < len(row) else ''
            amount_val = row[col_indices['amount']] if col_indices['amount'] < len(row) else ''
            owner_val = row[col_indices['owner']] if col_indices['owner'] < len(row) else ''
            receipt_val = row[col_indices['receipt']] if col_indices['receipt'] < len(row) else ''

            # Sanitize
            date_val = str(date_val).replace('\n', ' ').replace('\r', ' ') if date_val else ''
            detail_val = str(detail_val).replace('\n', ' ').replace('\r', ' ') if detail_val else ''
            type_val = str(type_val).replace('\n', ' ').replace('\r', ' ') if type_val else ''
            amount_val = str(amount_val).replace('\n', ' ').replace('\r', ' ') if amount_val else ''
            owner_val = str(owner_val).replace('\n', ' ').replace('\r', ' ') if owner_val else ''
            receipt_val = str(receipt_val).replace('\n', ' ').replace('\r', ' ') if receipt_val else ''

            # Skip total/summary rows (where date is missing or it's clearly a total)
            if not date_val.strip() or detail_val.lower().strip() in ['total', 'grand total', 'subtotal']:
                continue

            # Find matching receipt
            available_filenames = list(receipt_mapping.keys())
            matched_filename, score = fuzzy_match_filename(receipt_val, available_filenames)

            if matched_filename:
                mapping_info = receipt_mapping[matched_filename]
                reference_to_csv = f"{xlsx_path.name}:Row_{mapping_info['row_number']}"
                file_location = mapping_info['full_path']
            else:
                reference_to_csv = "NOT_FOUND"
                file_location = "NOT_FOUND"
                if receipt_val.strip():
                    print(f"⚠️  No match found for receipt: '{receipt_val}' in {xlsx_path.name}")

            # Create row with only required columns + extras
            row_data = [
                date_val, detail_val, type_val, amount_val, owner_val, receipt_val,
                f"{subdir_name}/{xlsx_path.name}", reference_to_csv, file_location
            ]

            rows.append(row_data)

    except Exception as e:
        print(f"Error processing {xlsx_path}: {e}", file=sys.stderr)

    return rows

def combine_excel_files():
    """Main function to combine all Excel files"""

    expense_dir = Path("Expense details")
    mapping_csv = "expense_receipt_mapping.csv"

    if not expense_dir.exists():
        print(f"Error: {expense_dir} directory not found")
        return False

    if not Path(mapping_csv).exists():
        print(f"Error: {mapping_csv} not found")
        return False

    # Load receipt mapping
    print("Loading receipt mapping...")
    receipt_mapping = load_receipt_mapping(mapping_csv)
    if not receipt_mapping:
        print("Failed to load receipt mapping")
        return False

    print(f"Loaded {len(receipt_mapping)} receipt mappings")

    # Define output CSV columns - only the required columns + 3 reference columns
    output_columns = [
        'Date', 'Detail', 'Type', 'Amount', 'Owner', 'Receipt',
        'XLS_REFERENCE', 'REFERENCE_TO_CSV', 'FILE_LOCATION'
    ]

    all_rows = []
    processed_files = 0
    expected_count = 0  # number of rows after headers across all excels

    # Process each event folder
    event_folders = sorted([f for f in expense_dir.iterdir() if f.is_dir()])

    for event_folder in event_folders:
        event_name = event_folder.name
        print(f"\n📁 Processing {event_name}")

        # Find nested folder
        nested_folder = find_nested_folder(event_folder)
        if not nested_folder:
            print(f"   ⚠️  No nested folder found")
            continue

        # Find all Excel files
        xlsx_files = find_all_xlsx_files(nested_folder)
        if not xlsx_files:
            print(f"   ⚠️  No Excel files found")
            continue

        # Process all Excel files
        for xlsx_file in xlsx_files:
            print(f"   📄 {xlsx_file.name}")
            rows = process_excel_file(xlsx_file, event_name, receipt_mapping)
            all_rows.extend(rows)
            processed_files += 1
            
            # Count expected rows (skip header)
            try:
                wb_tmp = load_workbook(xlsx_file)
                ws_tmp = wb_tmp.active
                # subtract header and any empty rows
                for r in ws_tmp.iter_rows(values_only=True):
                    if r and any(cell is not None for cell in r):
                        expected_count += 1
                expected_count -= 1
            except Exception:
                pass

            print(f"      ✓ {len(rows)} rows")

    # Write combined CSV
    if all_rows:
        output_csv = "combined_expense_data.csv"
        try:
            with open(output_csv, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(output_columns)
                writer.writerows(all_rows)

            print(f"\n{'='*80}")
            print(f"✅ Combined CSV created: {output_csv}")
            print(f"📊 Total rows generated : {len(all_rows)}")
            print(f"📊 Expected rows from excels: {expected_count}")
            if expected_count != len(all_rows):
                print("⚠️  Row count mismatch! Check for blank/hidden rows or parsing issues.")
            print(f"📁 Files processed: {processed_files}")
            print(f"{'='*80}\n")
            return True

        except Exception as e:
            print(f"Error writing combined CSV: {e}")
            return False
    else:
        print("❌ No data found to combine")
        return False

if __name__ == "__main__":
    success = combine_excel_files()
    sys.exit(0 if success else 1)
