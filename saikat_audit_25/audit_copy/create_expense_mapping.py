#!/usr/bin/env python3
"""
Create a flattened CSV that maps all receipts to their event and Excel file entries
"""

import os
import csv
from pathlib import Path
from openpyxl import load_workbook
import sys

def find_nested_folder(parent_path):
    """Find the nested folder with the same name pattern"""
    for item in parent_path.iterdir():
        if item.is_dir():
            return item
    return None

def find_xlsx_file(folder_path):
    """Find the Excel file in a folder"""
    for item in folder_path.iterdir():
        if item.is_file() and item.suffix.lower() == '.xlsx':
            return item
    return None

def find_receipts_folder(folder_path):
    """Find the Receipts folder (handles variations like 'receipt', 'Receipts', etc.)"""
    for item in folder_path.iterdir():
        if item.is_dir() and 'receipt' in item.name.lower():
            return item
    return None

def read_excel_entries(xlsx_path):
    """Read all non-empty rows from Excel file"""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        entries = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Skip empty rows
            if row and any(cell is not None for cell in row):
                # Create a string representation of the row
                row_data = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                entries.append({
                    'row_number': row_idx,
                    'content': row_data
                })
        return entries
    except Exception as e:
        print(f"Error reading Excel file {xlsx_path}: {e}", file=sys.stderr)
        return []

def get_receipts(receipts_folder):
    """Get all receipt files from a folder"""
    receipts = []
    try:
        for item in sorted(receipts_folder.iterdir()):
            if item.is_file():
                receipts.append(item)
    except Exception as e:
        print(f"Error reading receipts folder: {e}", file=sys.stderr)
    return receipts

def create_expense_receipt_mapping():
    """Main function to create the flattened CSV mapping"""
    
    expense_dir = Path("Expense details")
    if not expense_dir.exists():
        print(f"Error: {expense_dir} directory not found")
        return False
    
    output_csv = "expense_receipt_mapping.csv"
    all_rows = []
    
    event_folders = sorted([f for f in expense_dir.iterdir() if f.is_dir()])
    
    print(f"\nProcessing {len(event_folders)} event folders...\n")
    
    for event_folder in event_folders:
        event_name = event_folder.name
        print(f"📁 {event_name}")
        
        # Find nested folder
        nested_folder = find_nested_folder(event_folder)
        if not nested_folder:
            print(f"   ⚠️  No nested folder found")
            continue
        
        # Find Excel file
        xlsx_file = find_xlsx_file(nested_folder)
        if not xlsx_file:
            print(f"   ⚠️  No Excel file found")
            continue
        
        # Find Receipts folder
        receipts_folder = find_receipts_folder(nested_folder)
        if not receipts_folder:
            print(f"   ⚠️  No Receipts folder found")
            continue
        
        # Read Excel entries
        excel_entries = read_excel_entries(xlsx_file)
        print(f"   ✓ Excel file: {xlsx_file.name} ({len(excel_entries)} entries)")
        
        # Get receipts
        receipts = get_receipts(receipts_folder)
        print(f"   ✓ Found {len(receipts)} receipts")
        
        # Create rows for CSV
        for receipt_file in receipts:
            # Convert absolute path to relative path from current directory
            try:
                receipt_path = receipt_file.relative_to(Path.cwd())
            except ValueError:
                # If relative_to fails, just use the path as-is
                receipt_path = receipt_file
            
            row = {
                'Event': event_name,
                'Excel_Filename': xlsx_file.name,
                'Total_Excel_Entries': len(excel_entries),
                'Receipt_Filename': receipt_file.name,
                'Receipt_Full_Path': str(receipt_path),
                'Excel_Entry_Count': len(excel_entries)
            }
            all_rows.append(row)
        
        print()
    
    # Write CSV
    if all_rows:
        try:
            with open(output_csv, 'w', newline='', encoding='utf-8') as f:
                fieldnames = [
                    'Event',
                    'Excel_Filename',
                    'Total_Excel_Entries',
                    'Receipt_Filename',
                    'Receipt_Full_Path',
                    'Excel_Entry_Count'
                ]
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(all_rows)
            
            print(f"{'='*70}")
            print(f"✅ CSV file created: {output_csv}")
            print(f"📊 Total receipt entries: {len(all_rows)}")
            print(f"{'='*70}\n")
            return True
        except Exception as e:
            print(f"Error writing CSV file: {e}")
            return False
    else:
        print("❌ No receipt entries found")
        return False

if __name__ == "__main__":
    success = create_expense_receipt_mapping()
    sys.exit(0 if success else 1)
